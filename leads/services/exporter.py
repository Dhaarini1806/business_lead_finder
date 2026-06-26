"""
Export utilities for Forge OS — Lead Intelligence platform.

Generates polished ``leads.xlsx`` workbooks (OpenPyXL) and ``leads.csv``
files. Column sets adapt to the extraction source so each module exports the
fields that matter for it, while a unified column set is also available.
"""

from __future__ import annotations

import csv
import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (header, attribute) pairs per source. Callables receive the lead object.
COLUMN_SETS: dict[str, list[tuple[str, str]]] = {
    "google_maps": [
        ("Name", "name"), ("Category", "category"), ("Address", "address"),
        ("Phone", "phone"), ("Email", "email"), ("Website", "website"),
        ("Rating", "rating"), ("Reviews", "reviews"),
        ("Latitude", "latitude"), ("Longitude", "longitude"),
        ("Maps Link", "google_maps_link"),
    ],
    "linkedin": [
        ("Name", "name"), ("Position", "position"), ("Company", "company"),
        ("Industry", "industry"), ("Company Size", "company_size"),
        ("Email", "email"), ("LinkedIn URL", "linkedin_url"),
    ],
    "website": [
        ("Name", "name"), ("Email", "email"), ("Phone", "phone"),
        ("Website", "website"), ("Address", "address"),
        ("Category", "category"),
    ],
    "email_finder": [
        ("Name", "name"), ("Company", "company"), ("Email", "email"),
        ("Status", "email_status"), ("Confidence", "email_confidence"),
        ("Website", "website"),
    ],
    "enrichment": [
        ("Name", "name"), ("Industry", "industry"), ("Category", "category"),
        ("Website", "website"), ("Email", "email"), ("Phone", "phone"),
        ("Address", "address"),
    ],
}

DEFAULT_COLUMNS = COLUMN_SETS["google_maps"]

HEADER_FILL = "0E7490"  # cyan-700, matches Forge accent


def _columns_for(source: str | None) -> list[tuple[str, str]]:
    return COLUMN_SETS.get(source or "", DEFAULT_COLUMNS)


def _value(obj, attr):
    val = getattr(obj, attr, "")
    if val is None:
        return ""
    return val


def build_workbook(businesses: Iterable, source: str | None = None) -> Workbook:
    """Build a styled OpenPyXL workbook for the given leads."""
    columns = _columns_for(source)
    headers = [h for h, _ in columns]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center")

    sheet.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    link_cols = {
        i for i, (h, _) in enumerate(columns, start=1)
        if h in ("Website", "Maps Link", "LinkedIn URL")
    }

    for biz in businesses:
        sheet.append([_value(biz, attr) for _, attr in columns])

    for row in range(2, sheet.max_row + 1):
        for col in link_cols:
            cell = sheet.cell(row=row, column=col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0E7490", underline="single")

    for idx, (header, _) in enumerate(columns, start=1):
        width = 40 if header in ("Address", "Maps Link", "Website", "LinkedIn URL") else 20
        sheet.column_dimensions[get_column_letter(idx)].width = width

    sheet.freeze_panes = "A2"
    return workbook


def build_csv(businesses: Iterable, source: str | None = None) -> bytes:
    """Build a CSV byte string for the given leads."""
    columns = _columns_for(source)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([h for h, _ in columns])
    for biz in businesses:
        writer.writerow([_value(biz, attr) for _, attr in columns])
    return buffer.getvalue().encode("utf-8-sig")
